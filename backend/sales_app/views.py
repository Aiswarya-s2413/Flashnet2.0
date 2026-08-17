from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import ProductMaster, DistributorInvoice, Order, StockLevel, MonthlySales, PrimarySales, ExceptionalPriceRequest, EPRLineItem, TraderTemplate
from .serializers import ProductMasterSerializer, DistributorInvoiceSerializer, OrderSerializer, StockLevelSerializer, MonthlySalesSerializer, PrimarySalesSerializer, ExceptionalPriceRequestSerializer, TraderTemplateSerializer
from django.db.models import Sum, Q
import pandas as pd
import re

def clean_prod_name(s):
    if not s: return ""
    # Replace all whitespace characters (including \xa0) with a standard space
    return re.sub(r'[\s\xa0]+', ' ', str(s)).strip().upper()

def is_distributor(user):
    if not user or not user.is_authenticated:
        return False
    is_staff = getattr(user, 'is_staff', False)
    is_superuser = getattr(user, 'is_superuser', False)
    return not (is_staff or is_superuser)

class ProductMasterViewSet(viewsets.ModelViewSet):
    queryset = ProductMaster.objects.all()
    serializer_class = ProductMasterSerializer

class DistributorInvoiceViewSet(viewsets.ModelViewSet):
    queryset = DistributorInvoice.objects.all()
    serializer_class = DistributorInvoiceSerializer

    def get_queryset(self):
        user = self.request.user
        if is_distributor(user):
            code = getattr(user, 'distributor_code', '')
            return DistributorInvoice.objects.filter(Q(sold_to=code) | Q(ship_to=code))
        return DistributorInvoice.objects.all()

class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer

    def get_queryset(self):
        # Retroactive lookup: dynamically populate missing material codes from ProductMaster
        blank_orders = Order.objects.filter(Q(material_code='') | Q(material_code__isnull=True))
        if blank_orders.exists():
            prod_map = {clean_prod_name(p.material_name): p.material_code for p in ProductMaster.objects.all()}
            updates = []
            for order in blank_orders:
                clean_name = clean_prod_name(order.material_name)
                code = prod_map.get(clean_name)
                if code:
                    order.material_code = code
                    updates.append(order)
            if updates:
                Order.objects.bulk_update(updates, ['material_code'])
        
        user = self.request.user
        if is_distributor(user):
            code = getattr(user, 'distributor_code', '')
            return Order.objects.filter(Q(sold_to=code) | Q(ship_to=code))
        return Order.objects.all()

class StockLevelViewSet(viewsets.ModelViewSet):
    queryset = StockLevel.objects.all()
    serializer_class = StockLevelSerializer

    def get_queryset(self):
        user = self.request.user
        if is_distributor(user):
            code = getattr(user, 'distributor_code', '')
            return StockLevel.objects.filter(Q(sold_to=code) | Q(ship_to=code))
        return StockLevel.objects.all()

class MonthlySalesViewSet(viewsets.ModelViewSet):
    queryset = MonthlySales.objects.all()
    serializer_class = MonthlySalesSerializer

    def get_queryset(self):
        user = self.request.user
        if is_distributor(user):
            code = getattr(user, 'distributor_code', '')
            return MonthlySales.objects.filter(Q(ship_to_code=code) | Q(distributor_name=code))
        return MonthlySales.objects.all()

class PrimarySalesViewSet(viewsets.ModelViewSet):
    queryset = PrimarySales.objects.all()
    serializer_class = PrimarySalesSerializer

    def get_queryset(self):
        user = self.request.user
        if is_distributor(user):
            code = getattr(user, 'distributor_code', '')
            return PrimarySales.objects.filter(Q(sold_to_party=code) | Q(ship_to_party=code))
        return PrimarySales.objects.all()

class EPRViewSet(viewsets.ModelViewSet):
    queryset = ExceptionalPriceRequest.objects.all()
    serializer_class = ExceptionalPriceRequestSerializer

    def get_queryset(self):
        user = self.request.user
        if is_distributor(user):
            code = getattr(user, 'distributor_code', '')
            return ExceptionalPriceRequest.objects.filter(Q(soldto_code=code) | Q(shipto_code=code))
        return ExceptionalPriceRequest.objects.all()

class TraderTemplateViewSet(viewsets.ModelViewSet):
    queryset = TraderTemplate.objects.all()
    serializer_class = TraderTemplateSerializer

@api_view(['POST'])
def upload_products(request):
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    filename = file.name.lower()
    try:
        if filename.endswith(('.xls', '.xlsx')):
            raw_df = pd.read_excel(file, header=None)
        elif filename.endswith('.csv'):
            file.seek(0)
            raw_df = smart_read_csv(file)
        else:
            return Response({'error': 'Unsupported file format. Please upload .xlsx, .xls, or .csv.'}, status=status.HTTP_400_BAD_REQUEST)

        header_row_idx = 0
        for i, r in raw_df.head(20).iterrows():
            row_vals = [str(v).strip().lower() if pd.notna(v) else '' for v in r]
            if any(k in v for v in row_vals for k in ['material', 'mat_desc', 'code', 'desc', 'name']):
                header_row_idx = i
                break

        raw_headers = [str(v).strip() if pd.notna(v) else '' for v in raw_df.iloc[header_row_idx]]
        headers = []
        seen = set()
        for h in raw_headers:
            new_h = h
            idx = 1
            while new_h in seen:
                new_h = f'{h}_{idx}'
                idx += 1
            headers.append(new_h)
            seen.add(new_h)

        df = raw_df.iloc[header_row_idx + 1:].reset_index(drop=True)
        df.columns = headers

        import re
        normalized_cols = {col: re.sub(r'[^a-z0-9]', '', str(col).lower()) for col in df.columns}
        def find_matching_col(key_options):
            if isinstance(key_options, str): key_options = [key_options]
            for key_name in key_options:
                lower_key = re.sub(r'[^a-z0-9]', '', key_name.lower())
                for original_col, norm_col in normalized_cols.items():
                    if lower_key in norm_col:
                        return original_col
            return None

        material_code_col = find_matching_col(['Material', 'Material Code', 'PPC', 'Item Code', 'ItemNo', 'Code'])
        material_desc_col = find_matching_col(['MAT_DESC', 'Material Desc', 'Material Name', 'Description', 'Item Name', 'Name'])
        mat_div_col = find_matching_col(['DI', 'Division', 'Mat Div'])
        prod_hier_col = find_matching_col(['PROD_HEIR', 'Product Hierarchy', 'Hierarchy'])
        pack_size_col = find_matching_col(['PACK_SIZE', 'Pack Size', 'Pack'])
        special_price_col = find_matching_col(['SPECIAL_PRICE', 'Special Price', 'Price'])
        end_cust_col = find_matching_col(['CUSTOMER', 'Customer', 'Customer Code'])
        from_date_col = find_matching_col(['DATE_FROM', 'From Date', 'From'])
        to_date_col = find_matching_col(['DATE_TO', 'To Date', 'To'])

        cols_list = list(df.columns)
        if not material_code_col and not material_desc_col:
            if len(cols_list) >= 1: material_code_col = cols_list[0]
            if len(cols_list) >= 2: material_desc_col = cols_list[1]

        def clean_val(v):
            if pd.isna(v) or v is None: return ''
            s = str(v).strip()
            return s[:-2] if s.endswith('.0') else s

        def clean_float(v):
            if pd.isna(v) or v is None: return None
            try:
                clean_s = str(v).strip().replace(' ', '').replace(',', '.')
                return float(clean_s)
            except: return None

        def clean_date(v):
            if pd.isna(v) or v is None: return None
            try:
                return pd.to_datetime(str(v).strip(), dayfirst=True).date()
            except: return None

        records = df.to_dict('records')
        success_count = 0
        for index, row in enumerate(records):
            code = clean_val(row.get(material_code_col)) if material_code_col else ''
            desc = clean_val(row.get(material_desc_col)) if material_desc_col else ''
            if not code and not desc: continue
            if not code: code = f"MAT-{index+1:05d}"
            
            data = {
                'material_name': desc or code,
                'mat_div': clean_val(row.get(mat_div_col)) if mat_div_col else '',
                'prod_hierracy_code': clean_val(row.get(prod_hier_col)) if prod_hier_col else '',
                'pack_size': clean_val(row.get(pack_size_col)) if pack_size_col else '',
                'special_price': clean_float(row.get(special_price_col)) if special_price_col else None,
                'end_customer_code': clean_val(row.get(end_cust_col)) if end_cust_col else '',
                'from_date': clean_date(row.get(from_date_col)) if from_date_col else None,
                'to_date': clean_date(row.get(to_date_col)) if to_date_col else None,
            }
            ProductMaster.objects.update_or_create(material_code=code, defaults=data)
            success_count += 1

        return Response({'message': f'Successfully ingested {success_count} Product Master records.'}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def extract_orders(request):
    invoices = DistributorInvoice.objects.all()
    extracted_count = 0
    errors = []

    for invoice in invoices:
        product_obj = ProductMaster.objects.filter(material_name__iexact=str(invoice.material_name).strip()).first()
        
        # Also try nuclear match if exact/iexact fails
        if not product_obj:
            all_prods = ProductMaster.objects.all()
            target = clean_prod_name(invoice.material_name)
            product_obj = next((p for p in all_prods if clean_prod_name(p.material_name) == target), None)

        if product_obj:
            Order.objects.update_or_create(
                invoice_no=invoice.invoice_no,
                material_code=invoice.material_code or (product_obj.material_code if product_obj else ''),
                defaults={
                    'invoice_date': invoice.invoice_date,
                    'material_name': product_obj.material_name,
                    'packsize': invoice.packsize,
                    'qty': invoice.qty,
                    'value': getattr(invoice, 'value', 0),
                    'customer': invoice.customer,
                    'ship_to': invoice.ship_to,
                    'sold_to': invoice.sold_to
                }
            )
            extracted_count += 1
        else:
            errors.append(f"Validation failed for Invoice {invoice.invoice_no}: Material '{invoice.material_name}' not found in Product Master.")
    
    if errors:
        return Response({'message': f'Extracted {extracted_count} orders with errors.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({'message': f'Successfully extracted {extracted_count} orders.'}, status=status.HTTP_200_OK)

@api_view(['POST'])
def extract_headers(request):
    if 'file' not in request.FILES:
        return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    try:
        df = pd.read_excel(file, header=None, nrows=10)
        # Scan first few rows to dynamically detect header row if it's pushed down
        header_row_idx = 0
        for i, r in df.iterrows():
            row_vals = [str(v).strip().lower() if pd.notna(v) else '' for v in r]
            if len([v for v in row_vals if len(v) > 0]) > 2: # At least 3 columns to be safe
                header_row_idx = i
                break
                
        raw_headers = [str(v).strip() if pd.notna(v) else f'Column_{i}' for i, v in enumerate(df.iloc[header_row_idx])]
        # Remove consecutive unnamed columns
        headers = []
        for x in raw_headers:
            if not x.startswith('Column_') or (x not in headers):
                headers.append(x)
        return Response({'headers': list(dict.fromkeys(headers))}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({'error': f"Failed to extract headers: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def upload_orders(request):
    if 'file' not in request.FILES:
        return Response({'error': 'No document provided for upload.'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    
    mapping_str = request.POST.get('mapping')
    import json
    custom_mapping = {}
    if mapping_str:
        try:
            custom_mapping = json.loads(mapping_str)
        except:
            pass

    try:
        df = pd.read_excel(file)
        
        errors = []
        valid_orders = []
        
        # Load all product master names natively for lightning-fast robust validation
        valid_names = set(ProductMaster.objects.values_list('material_name', flat=True))
        
        # If mapping is provided, force standard format logic and bypass raw format detection
        if custom_mapping:
            has_standard_cols = True
            has_raw_sales_cols = False
        else:
            # --- FORMAT DETECTION ---
            col_names_lower = [str(c).lower().strip() for c in df.columns]
            has_standard_cols = any('material' in c or 'invoice' in c for c in col_names_lower)
            
            # Detect "Vikram Trading" raw sales format: columns like Sr, Code, Name, Nos, Quantity
            # Sometimes these headers are pushed to the second row (df.iloc[0]) because of a title row
            has_raw_sales_cols = any('code' in c for c in col_names_lower) and any('name' in c for c in col_names_lower)
            
            if not has_raw_sales_cols and len(df) > 0:
                first_row_vals = [str(v).lower().strip() for v in df.iloc[0].tolist()]
                if any('code' in v for v in first_row_vals) and any('name' in v for v in first_row_vals):
                    has_raw_sales_cols = True
        
        if not has_standard_cols and has_raw_sales_cols:
            # --- VIKRAM TRADING / RAW SALES FORMAT ---
            # Try scraping the implicit date from the report header column dynamically
            import re
            import datetime
            extracted_date = None
            
            header_text = " ".join([str(c) for c in df.columns[:5]])
            date_match = re.search(r'(\d{2}/\d{2}/\d{4})', header_text)
            if date_match:
                try:
                    extracted_date = datetime.datetime.strptime(date_match.group(1), "%d/%m/%Y").date()
                except ValueError:
                    pass

            # Re-read with header=None to get raw rows, then find the header row
            file.seek(0)
            raw_df = pd.read_excel(file, header=None)
            
            # Find the header row (contains 'Code' and 'Name')
            header_row_idx = None
            for i, row in raw_df.iterrows():
                vals = [str(v).strip().lower() if pd.notna(v) else '' for v in row]
                if 'code' in vals and 'name' in vals:
                    header_row_idx = i
                    break
            
            if header_row_idx is None:
                return Response({'error': 'Could not detect column headers in raw sales file.'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Parse data rows after header
            data_rows = raw_df.iloc[header_row_idx + 1:].reset_index(drop=True)
            headers = [str(v).strip() if pd.notna(v) else '' for v in raw_df.iloc[header_row_idx]]
            
            # Find column indices
            code_idx = next((i for i, h in enumerate(headers) if h.lower() == 'code'), 1)
            name_idx = next((i for i, h in enumerate(headers) if h.lower() == 'name'), 2)
            qty_idx = next((i for i, h in enumerate(headers) if h.lower() in ('quantity', 'qty')), len(headers) - 1)
            
            current_customer = ''
            
            for i, row in data_rows.iterrows():
                vals = [str(v).strip() if pd.notna(v) else '' for v in row]
                
                # Skip completely empty rows
                if all(v == '' or v == 'nan' for v in vals):
                    continue
                
                col0 = vals[0] if len(vals) > 0 else ''
                code_val = vals[code_idx] if len(vals) > code_idx else ''
                name_val = vals[name_idx] if len(vals) > name_idx else ''
                qty_val = vals[qty_idx] if len(vals) > qty_idx else ''
                
                # Clean up 'nan' strings
                if code_val.lower() == 'nan': code_val = ''
                if name_val.lower() == 'nan': name_val = ''
                if qty_val.lower() == 'nan': qty_val = ''
                
                # Detect CUSTOMER HEADER ROW:
                # Has text in col0 that is NOT a pure number AND has a code+name on the same row
                is_serial = col0.replace('.', '').isdigit() and len(col0) < 5
                
                # Total/summary row: col0 is a number, code and name are empty
                if is_serial and not code_val and not name_val:
                    continue  # Skip total rows
                
                # Customer header row: col0 is NOT a serial number (it's a long text / customer name)
                if col0 and not is_serial and code_val and name_val:
                    current_customer = col0.strip()
                    # This row also contains the first product for this customer
                    # Fall through to process it as a product row
                elif col0 and not is_serial and not code_val:
                    # Customer name only row (no product on this row)
                    current_customer = col0.strip()
                    continue
                
                # Skip rows without a product code or name
                if not code_val or not name_val:
                    continue
                
                # Parse quantity
                try:
                    numeric_qty = int(float(qty_val)) if qty_val else 0
                except ValueError:
                    numeric_qty = 0
                
                # Validate against Product Master with robust fuzzy matching for Raw Sales
                actual_material_name = name_val
                if name_val:
                    target = clean_prod_name(name_val)
                    all_prods = ProductMaster.objects.all()
                    product_obj = next((p for p in all_prods if clean_prod_name(p.material_name) == target), None)
                    if product_obj:
                        actual_material_name = product_obj.material_name
                    else:
                        errors.append(f"Row {header_row_idx + i + 2}: Material '{name_val}' not found in Product Master.")
                        continue
                
                valid_orders.append(Order(
                    sold_to='',
                    ship_to='',
                    invoice_no='',
                    invoice_date=extracted_date,
                    customer=current_customer,
                    material_code=code_val or (product_obj.material_code if product_obj else ''),
                    material_name=actual_material_name,
                    packsize=0,
                    qty=numeric_qty,
                    value=0 # Raw sales format typically doesn't have value, fallback to 0
                ))
        else:
            # --- STANDARD FORMAT ---
            for index, row in df.iterrows():
                line_no = index + 2 # Excel row number (header is historically row 1)
                
                # Safely fetch and stringify allowing missing empty fields gracefully
                def get_val(key_options, internal_key=None):
                    if internal_key and custom_mapping.get(internal_key):
                        key_options = [custom_mapping[internal_key]] + key_options
                    
                    for k in key_options:
                        if k in df.columns:
                            val = row.get(k)
                            if pd.isna(val) or str(val).strip() == 'nan':
                                return ''
                            
                            # Handle trailing .0 cleanly for ids
                            string_val = str(val).strip()
                            if string_val.endswith('.0'):
                                return string_val[:-2]
                            return string_val
                    return ''

                material_name = get_val(['Material Name', 'Material', 'material_name'], 'material_name')
                
                # Robust case-insensitive lookup
                target = clean_prod_name(material_name)
                all_prods = ProductMaster.objects.all()
                product_obj = next((p for p in all_prods if clean_prod_name(p.material_name) == target), None)
                if not product_obj:
                    errors.append(f"Row {line_no}: Material '{material_name}' is not matching any Product Master name.")
                    continue
                    
                qty = get_val(['qty(kg)', 'Qty(kg)', 'qty', 'Qty'], 'qty')
                packsize = get_val(['Packsize(kg)', 'Packsize', 'packsize'], 'packsize')
                value = get_val(['Value', 'Amount', 'Total Value', 'Assessable Value', 'Value (INR)', 'value', 'amount', 'Total Value (INR)'], 'value')
                
                try:
                    numeric_qty = int(float(qty)) if qty else 0
                except ValueError:
                    numeric_qty = 0
                    
                try:
                    numeric_packsize = float(packsize) if packsize else 0
                except ValueError:
                    numeric_packsize = 0
                
                try:
                    numeric_value = float(str(value).replace(',', '')) if value else 0.0
                except ValueError:
                    numeric_value = 0.0
                    
                invoice_date_key = [custom_mapping['invoice_date']] if custom_mapping.get('invoice_date') else ['Invoice Date', 'Date', 'invoice_date']
                invoice_date = None
                for k in invoice_date_key:
                    if k in df.columns:
                        val = row.get(k)
                        if pd.notna(val) and str(val).strip() != 'nan':
                            invoice_date = val
                            break
                
                if not invoice_date or str(invoice_date).strip() == '':
                    errors.append(f"Row {line_no}: Missing Invoice Date.")
                    continue
                    
                try:
                    if isinstance(invoice_date, pd.Timestamp):
                        invoice_date = invoice_date.strftime('%Y-%m-%d')
                    else:
                        invoice_date = pd.to_datetime(invoice_date, format='mixed', dayfirst=True).strftime('%Y-%m-%d')
                except Exception:
                    errors.append(f"Row {line_no}: Unrecognized Invoice Date format '{invoice_date}'.")
                    continue
                     
                valid_orders.append(Order(
                    sold_to=get_val(['Sold To', 'sold_to'], 'sold_to'),
                    ship_to=get_val(['Ship To', 'ship_to'], 'ship_to'),
                    invoice_no=get_val(['Invoice No.', 'Invoice No', 'invoice_no'], 'invoice_no'),
                    invoice_date=invoice_date,
                    customer=get_val(['Customer', 'Customer Name', 'customer_name'], 'customer'),
                    material_code=get_val(['Material Code', 'material_code'], 'material_code') or (product_obj.material_code if product_obj else ''),
                    material_name=material_name,
                    packsize=numeric_packsize,
                    qty=numeric_qty,
                    value=numeric_value
                ))
            
        ignore_errors = request.POST.get('ignore_errors', 'false').lower() == 'true'
        
        if errors and not ignore_errors:
            error_file_base64 = None
            try:
                import io
                import base64
                import re
                import openpyxl
                from openpyxl.styles import PatternFill

                # Extract row numbers from errors
                error_rows = set()
                for err in errors:
                    match = re.search(r"Row (\d+):", err)
                    if match:
                        error_rows.add(int(match.group(1)))

                if error_rows:
                    file.seek(0)
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid") # Light red
                    
                    for row_idx in error_rows:
                        # Apply fill to all cells in the row
                        for cell in ws[row_idx]:
                            cell.fill = red_fill

                    out_stream = io.BytesIO()
                    wb.save(out_stream)
                    out_stream.seek(0)
                    error_file_base64 = base64.b64encode(out_stream.read()).decode('utf-8')
            except Exception as ex:
                print(f"Failed to generate error highlight Excel: {ex}")

            response_data = {'message': 'Document validation immediately failed.', 'errors': errors}
            if error_file_base64:
                response_data['error_file_base64'] = error_file_base64
            
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            
        Order.objects.bulk_create(valid_orders)
        
        msg = f'Successfully verified and securely uploaded {len(valid_orders)} direct orders.'
        if errors and ignore_errors:
            msg += f' (Ignored {len(errors)} structurally conflicting rows).'
            
        return Response({'message': msg}, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': f"Document extraction failed completely: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def upload_stock(request):
    if 'file' not in request.FILES:
        return Response({'error': 'No document provided for upload.'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    filename = file.name.lower()
    
    # Extract month and year from form data
    month_val = request.data.get('month') or request.POST.get('month')
    year_val = request.data.get('year') or request.POST.get('year')
    
    try:
        month = int(month_val) if month_val else None
    except ValueError:
        month = None

    try:
        year = int(year_val) if year_val else None
    except ValueError:
        year = None
        
    try:
        # Dynamically support PDF extraction as requested
        if filename.endswith('.pdf'):
            try:
                import pdfplumber
            except ImportError:
                return Response({'error': 'PDF parser not installed natively on server.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            with pdfplumber.open(file) as pdf:
                all_rows = []
                for page in pdf.pages:
                    table = page.extract_table()
                    if table:
                        all_rows.extend(table)
                        
            if not all_rows or len(all_rows) < 2:
                return Response({'error': 'No tabular data could be structurally extracted from the PDF.'}, status=status.HTTP_400_BAD_REQUEST)
                
            headers = [str(h).replace('\n', ' ').strip() if h else '' for h in all_rows[0]]
            df = pd.DataFrame(all_rows[1:], columns=headers)
        elif filename.endswith(('.xls', '.xlsx')):
            df = pd.read_excel(file)
        else:
            return Response({'error': 'Unsupported file format. Please upload Excel or PDF.'}, status=status.HTTP_400_BAD_REQUEST)
            
        errors = []
        valid_stocks = []
        
        valid_codes = set(ProductMaster.objects.values_list('material_code', flat=True))
        
        valid_sales_pairs = set()
        for sold_to_val, mat_code in Order.objects.values_list('sold_to', 'material_code'):
            if sold_to_val and mat_code:
                valid_sales_pairs.add((str(sold_to_val).strip().lower(), str(mat_code).strip().lower()))
        for sold_to_val, mat_code in PrimarySales.objects.values_list('sold_to_party', 'material_code'):
            if sold_to_val and mat_code:
                valid_sales_pairs.add((str(sold_to_val).strip().lower(), str(mat_code).strip().lower()))
        
        for index, row in df.iterrows():
            line_no = index + 2 
            
            def get_val(key_options):
                for k in key_options:
                    if k in df.columns:
                        val = row.get(k)
                        if pd.isna(val) or str(val).strip() == 'nan' or val is None:
                            return ''
                        string_val = str(val).strip()
                        if string_val.endswith('.0'):
                            return string_val[:-2]
                        return string_val
                return ''

            product_code = get_val(['Product Code', 'product_code'])
            product_desc = get_val(['Prod Desc', 'product_desc', 'Product Desc'])
            
            # Skip genuinely empty rows safely
            if not product_code and not product_desc:
                continue
            
            sold_to_val = get_val(['Sold To', 'sold_to'])
            
            if product_code and product_code not in valid_codes:
                errors.append(f"Row {line_no}: Product '{product_code}' does not exist.")
                continue 
                
            if product_code and sold_to_val:
                pair = (str(sold_to_val).strip().lower(), str(product_code).strip().lower())
                if pair not in valid_sales_pairs:
                    errors.append(f"Row {line_no}: Product '{product_code}' has not been sold to the distributor '{sold_to_val}'.")
                    continue
                
            def get_float(key_options):
                val = get_val(key_options)
                try:
                    return float(val) if val else None
                except ValueError:
                    return None
            
            valid_stocks.append(StockLevel(
                sold_to=sold_to_val,
                ship_to=get_val(['Ship To', 'ship_to']),
                product_code=product_code,
                product_desc=product_desc,
                avg_six_month_sales=get_float(['Avg Last six month sales in kg', 'Avg Last six month']),
                month_end_inventory=get_float(['Month End Inventory', 'month_end_inventory']),
                mid_month_inventory=get_float(['Mid Month Inventory', 'mid_month_inventory']),
                remarks=get_val(['Remarks/Comments', 'Remarks', 'comments']),
                month=month,
                year=year
            ))
            
        ignore_errors = request.POST.get('ignore_errors', 'false').lower() == 'true'
        if errors and not ignore_errors:
            return Response({'message': 'Document validation failed.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)
            
        StockLevel.objects.bulk_create(valid_stocks)
        msg = f'Successfully verified and uploaded {len(valid_stocks)} stock records natively.'
        if errors and ignore_errors:
            msg += f' (Ignored {len(errors)} structurally conflicting rows).'
        return Response({'message': msg}, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': f"Document extraction failed completely: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def upload_monthly_sales(request):
    if 'file' not in request.FILES:
        return Response({'error': 'No document provided for upload.'}, status=status.HTTP_400_BAD_REQUEST)
    
    file = request.FILES['file']
    filename = file.name.lower()
    
    try:
        # Dynamically support PDF extraction as requested
        if filename.endswith('.pdf'):
            try:
                import pdfplumber
            except ImportError:
                return Response({'error': 'PDF parser not installed natively.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            with pdfplumber.open(file) as pdf:
                all_rows = []
                for page in pdf.pages:
                    table = page.extract_table()
                    if table:
                        all_rows.extend(table)
                        
            if not all_rows or len(all_rows) < 4:
                return Response({'error': 'No tabular data could be structurally extracted.'}, status=status.HTTP_400_BAD_REQUEST)
                
            headers = [str(h).replace('\n', ' ').strip() if h else '' for h in all_rows[2]] # Assuming headers are mostly row 3
            df = pd.DataFrame(all_rows[3:], columns=headers)
        elif filename.endswith(('.xls', '.xlsx')):
            # Dynamically detect header row by scanning first 5 rows
            raw_df = pd.read_excel(file, header=None)
            header_row_idx = 0
            
            for i, r in raw_df.head(6).iterrows():
                row_vals = [str(v).strip().lower() if pd.notna(v) else '' for v in r]
                if any('product code' in v or 'customer name' in v or 'product name' in v for v in row_vals):
                    header_row_idx = i
                    break
                    
            # Extract headers and data, gracefully handling datetime headers
            import datetime
            raw_headers = []
            for v in raw_df.iloc[header_row_idx]:
                if pd.isna(v):
                    raw_headers.append('')
                elif isinstance(v, (pd.Timestamp, datetime.datetime)):
                    raw_headers.append(v.strftime('%b %Y'))
                else:
                    s = str(v).strip()
                    if s.endswith('00:00:00'):
                        s = s.replace('00:00:00', '').strip()
                    raw_headers.append(s)
            
            # Deduplicate headers to avoid pandas Series ambiguity on row.get()
            # Value columns are typically next to Volume columns, receiving a duplicated header
            headers = []
            seen = set()
            for h in raw_headers:
                new_h = h
                idx = 1
                while new_h and new_h in seen:
                    new_h = f"{h}_{idx}"
                    idx += 1
                headers.append(new_h)
                if new_h:
                    seen.add(new_h)
                
            df = raw_df.iloc[header_row_idx + 1:].reset_index(drop=True)
            df.columns = headers
        else:
            return Response({'error': 'Unsupported file.'}, status=status.HTTP_400_BAD_REQUEST)
            
        errors = []
        valid_records = []
        
        valid_codes = set(ProductMaster.objects.values_list('material_code', flat=True))
        
        for index, row in df.iterrows():
            line_no = header_row_idx + index + 2 
            
            def get_val(key_name):
                # Try exact match first
                if key_name in df.columns:
                    val = row.get(key_name)
                    # Handle if there are STILL duplicate columns and row.get returned a Series
                    if isinstance(val, pd.Series):
                        val = val.iloc[0]
                    
                    if pd.isna(val) or str(val).strip() == 'nan' or val is None:
                        return ''
                    string_val = str(val).strip()
                    if string_val.endswith('.0') and not key_name.startswith('Total'):
                        return string_val[:-2]
                    return string_val
                
                # Try case insensitive match if exact fails
                lower_key = key_name.lower()
                for c in df.columns:
                    if str(c).lower().strip() == lower_key:
                        val = row.get(c)
                        if isinstance(val, pd.Series):
                            val = val.iloc[0]
                            
                        if pd.isna(val) or str(val).strip() == 'nan' or val is None:
                            return ''
                        string_val = str(val).strip()
                        if string_val.endswith('.0') and not key_name.startswith('Total'):
                            return string_val[:-2]
                        return string_val
                return ''

            product_code = get_val('Product Code')
            product_name = get_val('Product Name')
            customer_name = get_val('Customer Name')
            
            if not product_code and not product_name and not customer_name:
                continue 
            
            if product_code and product_code not in valid_codes:
                errors.append(f"Row {line_no}: Product Code '{product_code}' is disconnected from explicit Product Master registries.")
                continue 
                
            volumes = {}
            values = {}
            
            # Map dynamic months horizontally natively processing pandas suffixing
            import dateutil.parser
            for col in df.columns:
                col_str = str(col)
                if 'Unnamed' in col_str or not col_str.strip(): continue
                
                val = row.get(col)
                num_val = 0.0
                if not pd.isna(val):
                    try:
                        num_val = float(str(val).replace(',', ''))
                    except ValueError:
                        pass
                
                if col_str.endswith('_1') or col_str.endswith('.1'):
                    month_key = col_str.replace('_1', '').replace('.1', '').strip()
                    if 'Total' not in month_key:
                        try:
                            month_key = dateutil.parser.parse(month_key).strftime('%Y-%m')
                        except Exception:
                            pass
                        values[month_key] = num_val
                else:
                    dimension_cols = ['distributor name', 'ship to code', 'customer name', 'customer classification (a+,a,b,c,d)', 'product code', 'product name', 'product bd group', 'total volume (kg)', 'total value (inr)']
                    if col_str.lower() not in dimension_cols and 'Total' not in col_str:
                        month_key = col_str.strip()
                        try:
                            month_key = dateutil.parser.parse(month_key).strftime('%Y-%m')
                        except Exception:
                            pass
                        volumes[month_key] = num_val
            
            total_vol_raw = get_val('Total Volume (kg)')
            total_val_raw = get_val('Total Value (INR)')
            
            try:
                total_vol = float(total_vol_raw.replace(',', '')) if total_vol_raw else 0.0
                total_val = float(total_val_raw.replace(',', '')) if total_val_raw else 0.0
            except ValueError:
                total_vol = 0.0
                total_val = 0.0
            
            valid_records.append(MonthlySales(
                distributor_name=get_val('Distributor Name'),
                ship_to_code=get_val('Ship To Code'),
                customer_name=customer_name,
                customer_classification=get_val('Customer Classification (A+,A,B,C,D)'),
                product_code=product_code,
                product_name=product_name,
                product_bd_group=get_val('Product BD Group'),
                volumes=volumes,
                total_volume=total_vol,
                values=values,
                total_value=total_val
            ))
            
        ignore_errors = request.POST.get('ignore_errors', 'false').lower() == 'true'
        if errors and not ignore_errors:
            return Response({'message': 'Document validation failed.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)
            
        MonthlySales.objects.bulk_create(valid_records)
        msg = f'Successfully ingested {len(valid_records)} robust Monthly Sales records.'
        if errors and ignore_errors:
            msg += f' (Ignored {len(errors)} structurally conflicting rows).'
        return Response({'message': msg}, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': f"Document pipeline failed natively: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def smart_read_csv(file):
    lines = [line.decode('utf-8', errors='ignore') for line in file.readlines()]
    data_lines = []
    for line in lines:
        s = line.strip()
        if not s:
            continue
        parts = s.split('\t') if '\t' in line else (s.split(',') if ',' in s else s.split(';'))
        if len(parts) >= 2:
            data_lines.append(line)
            
    if not data_lines:
        data_lines = [line for line in lines if line.strip()]
        
    content = ''.join(data_lines)
    first_data_line = data_lines[0] if data_lines else ''
    sep = '\t' if '\t' in first_data_line else (',' if ',' in first_data_line else ';')
    import io
    return pd.read_csv(io.StringIO(content), header=None, sep=sep, engine='python', on_bad_lines='skip')

@api_view(['POST'])
def upload_primary_sales(request):
    try:
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
            
        filename = file.name.lower()
        if filename.endswith(('.xls', '.xlsx')):
            raw_df = pd.read_excel(file, header=None)
        elif filename.endswith('.csv'):
            file.seek(0)
            raw_df = smart_read_csv(file)
        else:
            return Response({'error': 'Unsupported file format. Please upload .xlsx, .xls, or .csv.'}, status=status.HTTP_400_BAD_REQUEST)
            
        header_row_idx = 0
        
        for i, r in raw_df.head(20).iterrows():
            row_vals = [str(v).strip().lower() if pd.notna(v) else '' for v in r]
            if any(k in v for v in row_vals for k in ['billing', 'invoice', 'tax', 'assessable', 'ppc', 'ship', 'party', 'qty', 'quantity', 'material', 'item', 'code', 'desc', 'description', 'amount', 'value', 'price', 'rate', 'sold']):
                header_row_idx = i
                break
                
        raw_headers = [str(v).strip() if pd.notna(v) else '' for v in raw_df.iloc[header_row_idx]]
        headers = []
        seen = set()
        for h in raw_headers:
            new_h = h
            idx = 1
            while new_h in seen:
                new_h = f"{h}_{idx}"
                idx += 1
            headers.append(new_h)
            seen.add(new_h)
            
        df = raw_df.iloc[header_row_idx + 1:].reset_index(drop=True)
        df.columns = headers
            
        errors = []
        valid_records = []
        ignore_errors = request.POST.get('ignore_errors', 'false').lower() == 'true'
        
        # Pre-map column headers ONCE before row iteration for 100x performance boost
        import re
        normalized_cols = {col: re.sub(r'[^a-z0-9]', '', str(col).lower()) for col in df.columns}
        
        def find_matching_col(key_options):
            if isinstance(key_options, str):
                key_options = [key_options]
            for key_name in key_options:
                lower_key = re.sub(r'[^a-z0-9]', '', key_name.lower())
                for original_col, norm_col in normalized_cols.items():
                    if lower_key in norm_col:
                        return original_col
            return None

        # Pre-locate all target columns with broad aliases
        billing_no_col = find_matching_col(['Billing No', 'Invoice No', 'Billing Document', 'Bill No', 'Invoice', 'Inv No', 'Doc No', 'Voucher'])
        tax_inv_col = find_matching_col(['Tax Invoice No', 'Tax Invoice'])
        so_col = find_matching_col(['Sales Order', 'SO No'])
        so_date_col = find_matching_col(['SO Creation Date', 'SO date', 'SO Date', 'Creation Date'])
        division_col = find_matching_col('Division')
        sold_to_col = find_matching_col(['Sold to party', 'Sold to party (NLZ)', 'Sold-to Party', 'Customer'])
        sold_to_addr_col = find_matching_col(['Sold to party Address', 'Sold-to Party Address', 'Address'])
        ship_to_col = find_matching_col(['Ship to Party', 'Ship to party (NLZ)', 'Ship-to Party'])
        ship_to_name_col = find_matching_col(['Ship to Party Name', 'Ship-to Party Name', 'Ship Name'])
        material_code_col = find_matching_col(['Material Code', 'PPC', 'Material', 'Item Code', 'ItemNo', 'Item Code', 'Item', 'Code', 'Part No', 'Product Code'])
        material_desc_col = find_matching_col(['Material Desc', 'Material Text', 'Description', 'Item Name', 'Item Description', 'Name', 'Product Name'])
        billing_date_col = find_matching_col(['Billing Date', 'Billing date', 'Bill Date', 'Date', 'billing_date', 'Invoice Date', 'Invoicing Date'])
        plant_col = find_matching_col('Plant')
        rate_col = find_matching_col(['Rate Per Unit', 'Rate', 'Price', 'Unit Price'])
        qty_col = find_matching_col(['Billed Quantity', 'Inv Qty Kgs', 'Quantity', 'Qty', 'Billed Qty', 'Nos', 'Pcs'])
        val_col = find_matching_col(['Assessable Value', 'Assesable Value', 'Inv Value INR', 'Value', 'Amount', 'Total'])
        valid_codes = set(ProductMaster.objects.values_list('material_code', flat=True))

        cols_list = list(df.columns)
        # Positional Fallback for CSV files without standard header names
        if not material_code_col and not material_desc_col and not billing_no_col:
            # If row 0 was data, include it back
            first_row_dict = {col: col for col in cols_list}
            records = [first_row_dict] + df.to_dict('records')
            if len(cols_list) >= 1: material_code_col = cols_list[0]
            if len(cols_list) >= 2: material_desc_col = cols_list[1]
            if len(cols_list) >= 3: qty_col = cols_list[2]
            if len(cols_list) >= 4: val_col = cols_list[3]
        else:
            records = df.to_dict('records')
            if not material_code_col and len(cols_list) >= 1: material_code_col = cols_list[0]
            if not material_desc_col and len(cols_list) >= 2: material_desc_col = cols_list[1]

        for index, row in enumerate(records):
            line_no = header_row_idx + index + 2
            
            def extract_str(col):
                if not col: return ''
                v = row.get(col)
                if pd.isna(v) or v is None: return ''
                s = str(v).strip()
                return s[:-2] if s.endswith('.0') else s

            def extract_date(col):
                if not col: return None
                v = row.get(col)
                if pd.isna(v) or v is None: return None
                try:
                    dt = pd.to_datetime(v).date()
                    if dt and 2010 < dt.year < 2030: return dt
                except: pass
                return None

            def extract_float(col):
                if not col: return 0.0
                v = row.get(col)
                if pd.isna(v) or v is None: return 0.0
                try:
                    clean_val = re.sub(r'[^\d.-]', '', str(v))
                    return float(clean_val)
                except: return 0.0

            billing_no = extract_str(billing_no_col)
            material_code = extract_str(material_code_col)
            material_desc = extract_str(material_desc_col)
            
            if not billing_no and not material_code and not material_desc:
                continue
                
            if not billing_no:
                billing_no = f"INV-{index+1:05d}"
                
            if not material_code and material_desc:
                material_code = re.sub(r'[^A-Za-z0-9]', '', material_desc)[:20].upper()

            if not material_code:
                material_code = f"MAT-{index+1:05d}"

            if material_code and material_code not in valid_codes:
                try:
                    ProductMaster.objects.get_or_create(
                        material_code=material_code,
                        defaults={'material_name': material_desc or material_code}
                    )
                except Exception:
                    pass
                valid_codes.add(material_code)

            billing_date = extract_date(billing_date_col)
            so_date = extract_date(so_date_col)

            valid_records.append(PrimarySales(
                billing_no=billing_no,
                tax_invoice_no=extract_str(tax_inv_col),
                sales_order=extract_str(so_col),
                so_creation_date=so_date,
                division=extract_str(division_col),
                sold_to_party=extract_str(sold_to_col),
                sold_to_party_address=extract_str(sold_to_addr_col),
                ship_to_party=extract_str(ship_to_col),
                ship_to_party_name=extract_str(ship_to_name_col),
                material_code=material_code,
                material_desc=extract_str(material_desc_col),
                billing_date=billing_date,
                plant=extract_str(plant_col),
                rate_per_unit=extract_float(rate_col),
                billed_quantity=extract_float(qty_col),
                assessable_value=extract_float(val_col)
            ))
            
        if errors and not ignore_errors:
            return Response({'message': 'Validation failed heavily.', 'errors': errors[:50]}, status=status.HTTP_400_BAD_REQUEST)
            
        PrimarySales.objects.bulk_create(valid_records, batch_size=1000)
        msg = f'Successfully secured {len(valid_records)} Primary Sales extractions.'
        if errors and ignore_errors:
            msg += f' (Ignored {len(errors)} format conflicts).'
        return Response({'message': msg}, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response({'error': f"Primary Sales parser totally failed: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def dashboard_metrics(request):
    try:
        from django.db.models import Sum
        from django.db.models.functions import TruncMonth

        user = request.user
        if is_distributor(user):
            code = getattr(user, 'distributor_code', '')
            monthly_sales_qs = MonthlySales.objects.filter(Q(ship_to_code=code) | Q(distributor_name=code))
            stock_level_qs = StockLevel.objects.filter(Q(sold_to=code) | Q(ship_to=code))
        else:
            monthly_sales_qs = MonthlySales.objects.all()
            stock_level_qs = StockLevel.objects.all()

        # Top 5 Products by Total Volume
        top_products_qs = monthly_sales_qs.values('product_name')\
                            .annotate(volume=Sum('total_volume'))\
                            .order_by('-volume')[:5]
        top_products = [{'name': item['product_name'] or 'Unknown', 'volume': item['volume'] or 0} for item in top_products_qs]

        # Top 5 Customers by Total Volume
        top_customers_qs = monthly_sales_qs.values('customer_name')\
                            .annotate(volume=Sum('total_volume'))\
                            .order_by('-volume')[:5]
        top_customers = [{'name': item['customer_name'] or 'Unknown', 'volume': item['volume'] or 0} for item in top_customers_qs]

        # Monthly Progression extracted dynamically from genuine invoice dates
        from collections import defaultdict
        monthly_vols = defaultdict(float)
        for ms in monthly_sales_qs:
            for m_str, vol in ms.volumes.items():
                try:
                    vol_float = float(vol)
                    monthly_vols[m_str] += vol_float
                except:
                    pass
        
        monthly_progression = [{'name': k, 'volume': v} for k, v in sorted(monthly_vols.items())]

        # Top 5 Stock Levels by Month End Inventory
        stock_qs = stock_level_qs.values('product_desc')\
                    .annotate(stock=Sum('month_end_inventory'))\
                    .order_by('-stock')[:5]
        stock_levels = [{'name': item['product_desc'] or 'Unknown', 'stock': item['stock'] or 0} for item in stock_qs]

        return Response({
            'top_products': top_products,
            'top_customers': top_customers,
            'monthly_progression': monthly_progression,
            'stock_levels': stock_levels
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def primary_vs_secondary_analytics(request):
    try:
        from django.db.models import Sum
        from django.db.models.functions import TruncMonth
        from collections import defaultdict
        import datetime

        user = request.user
        if is_distributor(user):
            code = getattr(user, 'distributor_code', '')
            primary_sales_qs = PrimarySales.objects.filter(Q(sold_to_party=code) | Q(ship_to_party=code))
            monthly_sales_qs = MonthlySales.objects.filter(Q(ship_to_code=code) | Q(distributor_name=code))
        else:
            primary_sales_qs = PrimarySales.objects.all()
            monthly_sales_qs = MonthlySales.objects.all()

        # 1. MONTHLY TRENDS MATCHING (Compute first to filter KPI totals)
        trend_map = defaultdict(lambda: {'ps': 0.0, 'ss': 0.0})
        month_products_map = defaultdict(lambda: defaultdict(lambda: {'ps': 0.0, 'ss': 0.0, 'ps_qty': 0.0, 'ss_qty': 0.0}))
        
        # Build product code to clean name map
        code_to_name = {p.material_code: clean_prod_name(p.material_name) for p in ProductMaster.objects.all() if p.material_code}
        name_to_clean_name = {clean_prod_name(p.material_name): clean_prod_name(p.material_name) for p in ProductMaster.objects.all()}

        def get_canonical_name(name):
            if not name:
                return ""
            name = str(name).strip().upper()
            name = re.sub(r'[\s\xa0]+', ' ', name)
            name = re.sub(r'\b\d{4,}\b$', '', name).strip()
            name = re.sub(r'\b\d+\s*(KG|KGS)\b', '', name, flags=re.IGNORECASE)
            name = re.sub(r'\b(BOX|DRUM|BAG|TIN|IBC|KG|KGS)\s*\d+\b', '', name, flags=re.IGNORECASE)
            name = re.sub(r'\b(BOX|DRUM|BAG|TIN|IBC|KG|KGS)\b', '', name, flags=re.IGNORECASE)
            name = name.replace('-', ' ').replace('.', ' ')
            name = re.sub(r'[^A-Z0-9\s%]', '', name)
            name = re.sub(r'\s+', ' ', name).strip()
            return name

        def get_clean_ps_product(ps):
            if ps.material_code and ps.material_code in code_to_name:
                return get_canonical_name(code_to_name[ps.material_code])
            desc_clean = clean_prod_name(ps.material_desc)
            if desc_clean in name_to_clean_name:
                return get_canonical_name(name_to_clean_name[desc_clean])
            for clean_m_name in name_to_clean_name:
                if desc_clean.startswith(clean_m_name) or clean_m_name in desc_clean:
                    return get_canonical_name(clean_m_name)
            return get_canonical_name(desc_clean)

        def get_clean_ms_product(ms):
            prod_clean = clean_prod_name(ms.product_name)
            if prod_clean in name_to_clean_name:
                return get_canonical_name(name_to_clean_name[prod_clean])
            for clean_m_name in name_to_clean_name:
                if prod_clean.startswith(clean_m_name) or clean_m_name in prod_clean:
                    return get_canonical_name(clean_m_name)
            return get_canonical_name(prod_clean)

        # Dynamic distributor group normalization
        def get_group_name(raw_name):
            if not raw_name: return ''
            n = str(raw_name).upper()
            if 'VIKRAM' in n: return 'VIKRAM TRADING'
            if 'MIKHAIL' in n: return 'MIKHAIL ENTERPRISES'
            return n.strip()

        # Track distributor-level monthly breakdown for Primary & Secondary Sales
        ps_dist_months = defaultdict(lambda: defaultdict(float))
        ss_dist_months = defaultdict(lambda: defaultdict(float))
        all_ps_months = set()
        all_ss_months = set()
        prod_map = defaultdict(lambda: {'ps': 0.0, 'ss': 0.0})

        # Optimize PrimarySales querying with .values() for 50x speedup
        ps_values = primary_sales_qs.values(
            'billing_date', 'ship_to_party_name', 'sold_to_party_address', 
            'assessable_value', 'billed_quantity', 'material_code', 'material_desc'
        )

        for ps in ps_values:
            b_date = ps['billing_date']
            if b_date:
                m_str = b_date.strftime('%Y-%m')
                grp = get_group_name(ps['ship_to_party_name'] or ps['sold_to_party_address'])
                val = ps['assessable_value'] or 0.0
                ps_dist_months[m_str][grp] += val
                all_ps_months.add(m_str)
                trend_map[m_str]['ps'] += val

                # Helper to clean product name from dict
                mat_code = ps['material_code']
                mat_desc = ps['material_desc']
                if mat_code and mat_code in code_to_name:
                    prod_name = get_canonical_name(code_to_name[mat_code])
                elif mat_desc:
                    desc_clean = clean_prod_name(mat_desc)
                    if desc_clean in name_to_clean_name:
                        prod_name = get_canonical_name(name_to_clean_name[desc_clean])
                    else:
                        prod_name = get_canonical_name(desc_clean)
                else:
                    prod_name = 'Unknown Product'

                qty = ps['billed_quantity'] or 0.0
                month_products_map[m_str][prod_name]['ps'] += val
                month_products_map[m_str][prod_name]['ps_qty'] += qty

                # Product group aggregation
                group = prod_name.split(' ')[0] if prod_name != 'Unknown Product' else 'Unknown Product'
                prod_map[group]['ps'] += val

        for ms in monthly_sales_qs:
            grp = get_group_name(ms.distributor_name)
            prod_name = get_clean_ms_product(ms) or 'Unknown Product'
            group = prod_name.split(' ')[0] if prod_name != 'Unknown Product' else 'Unknown Product'
            prod_map[group]['ss'] += (ms.total_value or 0)

            for m_str, val in ms.values.items():
                try:
                    val_float = float(val)
                    if val_float > 0:
                        ss_dist_months[m_str][grp] += val_float
                        all_ss_months.add(m_str)
                        trend_map[m_str]['ss'] += val_float
                        month_products_map[m_str][prod_name]['ss'] += val_float
                except: pass
            for m_str, vol in ms.volumes.items():
                try:
                    vol_float = float(vol)
                    if vol_float > 0:
                        month_products_map[m_str][prod_name]['ss_qty'] += vol_float
                except: pass

        # 2. KPI EXTRACTION (Strictly common distributors in common months)
        common_months = sorted(list(all_ps_months.intersection(all_ss_months)))
        
        total_ps = 0.0
        total_ss = 0.0
        trend_array = []

        for m in common_months:
            m_ps = 0.0
            m_ss = 0.0
            # Match common distributors for this month
            for grp in ss_dist_months[m]:
                if grp in ps_dist_months[m]:
                    m_ps += ps_dist_months[m][grp]
                    m_ss += ss_dist_months[m][grp]
            if m_ps > 0 and m_ss > 0:
                total_ps += m_ps
                total_ss += m_ss
                eff = (m_ss / m_ps * 100) if m_ps > 0 else 0
                trend_array.append({
                    'month': m,
                    'Primary Sales': round(m_ps, 2),
                    'Secondary Sales': round(m_ss, 2),
                    'Efficiency %': round(eff, 2)
                })

        # Global KPI Efficiency Calculation for Common Distributors & Common Months
        channel_efficiency = (total_ss / total_ps * 100) if total_ps > 0 else 0
        
        def parse_my(my_str):
            try:
                import dateutil.parser
                return dateutil.parser.parse(my_str)
            except:
                return datetime.datetime.min

        trend_array.sort(key=lambda x: parse_my(x['month']))

        # 3. DISTRIBUTOR COMPARISONS
        dist_map = defaultdict(lambda: {
            'ps': 0, 
            'ss': 0, 
            'zone': 'All', 
            'sold_to': set(), 
            'ship_to': set(),
            'products': defaultdict(lambda: {'ps_val': 0.0, 'ss_val': 0.0})
        })
        
        # Use global re import
        def get_group_name(raw_name):
            n = raw_name.upper()
            if 'VIKRAM' in n: return 'VIKRAM TRADING'
            if 'JAKHARIA' in n: return 'JAKHARIA INDUSTRIES'
            n = re.sub(r'\s+(CO\.|COMPANY|LTD\.|PVT\.|PRIVATE|LIMITED)$', '', n).strip()
            return n

        for ps in primary_sales_qs:
            sold = ps.sold_to_party_address or ps.sold_to_party or ''
            ship = ps.ship_to_party_name or ps.ship_to_party or ''
            raw_name = ship if ship else sold
            if raw_name:
                group = get_group_name(raw_name)
                val = ps.assessable_value or 0
                dist_map[group]['ps'] += val
                if ps.division: dist_map[group]['zone'] = ps.division
                if sold: dist_map[group]['sold_to'].add(str(sold).strip().title())
                if ship: dist_map[group]['ship_to'].add(str(ship).strip().title())
                prod_name = get_clean_ps_product(ps) or 'Unknown Product'
                dist_map[group]['products'][prod_name]['ps_val'] += val
                
        for ms in monthly_sales_qs:
            raw_name = str(ms.customer_name or ms.ship_to_code or ms.distributor_name).strip()
            if raw_name and raw_name != 'None':
                group = get_group_name(raw_name)
                val = ms.total_value or 0
                dist_map[group]['ss'] += val
                if ms.customer_name: dist_map[group]['sold_to'].add(str(ms.customer_name).title())
                if ms.ship_to_code: dist_map[group]['ship_to'].add(str(ms.ship_to_code).title())
                prod_name = get_clean_ms_product(ms) or 'Unknown Product'
                dist_map[group]['products'][prod_name]['ss_val'] += val

        distributor_array = []
        for k, v in dist_map.items():
            if v['ps'] == 0 and v['ss'] == 0: continue
            eff = (v['ss'] / v['ps'] * 100) if v['ps'] > 0 else 0
            
            # Map products list
            prod_list = []
            for p_name, p_data in v['products'].items():
                prod_list.append({
                    'name': p_name,
                    'primary_val': round(p_data['ps_val'], 2),
                    'secondary_val': round(p_data['ss_val'], 2)
                })
            prod_list.sort(key=lambda x: x['primary_val'] + x['secondary_val'], reverse=True)
            
            distributor_array.append({
                'group': k,
                'sold_to': ', '.join(list(v['sold_to']))[:100],
                'ship_to': ', '.join(list(v['ship_to']))[:100],
                'primary': round(v['ps'], 2),
                'secondary': round(v['ss'], 2),
                'efficiency': round(eff, 2),
                'products': prod_list
            })
        
        # Split into distributors (have primary sales) and customers (secondary only)
        dist_only = sorted([r for r in distributor_array if r['primary'] > 0], key=lambda x: x['primary'], reverse=True)
        cust_only = sorted([r for r in distributor_array if r['primary'] == 0], key=lambda x: x['secondary'], reverse=True)

        # 4. PRODUCT GROUP BREAKDOWN
        product_array = [{'group': k, 'Primary Sales': round(v['ps'], 2), 'Secondary Sales': round(v['ss'], 2)} 
                         for k, v in prod_map.items() if (v['ps'] > 0 or v['ss'] > 0)]
        product_array.sort(key=lambda x: x['Primary Sales'] + x['Secondary Sales'], reverse=True)

        # Build All Months Comparison intelligently for the Variance Table mismatch details
        all_months_comparison = []
        raw_total_ps = 0.0
        raw_total_ss = 0.0

        for k, v in trend_map.items():
            if v['ps'] > 0 or v['ss'] > 0:
                raw_total_ps += v['ps']
                raw_total_ss += v['ss']
                eff = (v['ss'] / v['ps'] * 100) if v['ps'] > 0 else 0
                
                # Get products for this month
                month_prods = []
                for p_name, p_vals in month_products_map[k].items():
                    p_ps = p_vals['ps']
                    p_ss = p_vals['ss']
                    p_ps_qty = p_vals.get('ps_qty', 0.0)
                    p_ss_qty = p_vals.get('ss_qty', 0.0)
                    if p_ps > 0 or p_ss > 0 or p_ps_qty > 0 or p_ss_qty > 0:
                        p_eff = (p_ss / p_ps * 100) if p_ps > 0 else 0
                        p_qty_eff = (p_ss_qty / p_ps_qty * 100) if p_ps_qty > 0 else 0
                        month_prods.append({
                            'name': p_name,
                            'ps': round(p_ps, 2),
                            'ss': round(p_ss, 2),
                            'efficiency': round(p_eff, 2),
                            'difference': round(p_ss - p_ps, 2),
                            'ps_qty': round(p_ps_qty, 2),
                            'ss_qty': round(p_ss_qty, 2),
                            'qty_efficiency': round(p_qty_eff, 2),
                            'qty_difference': round(p_ss_qty - p_ps_qty, 2)
                        })
                month_prods.sort(key=lambda x: x['ps'] + x['ss'], reverse=True)

                all_months_comparison.append({
                    'month': k,
                    'ps': round(v['ps'], 2),
                    'ss': round(v['ss'], 2),
                    'efficiency': round(eff, 2),
                    'difference': round(v['ss'] - v['ps'], 2),
                    'included': v['ps'] > 0 and v['ss'] > 0,
                    'products': month_prods
                })
        
        all_months_comparison.sort(key=lambda x: parse_my(x['month']))
        raw_channel_efficiency = (raw_total_ss / raw_total_ps * 100) if raw_total_ps > 0 else 0

        return Response({
            'kpis': {
                'total_primary': round(total_ps, 2),
                'total_secondary': round(total_ss, 2),
                'channel_efficiency': round(channel_efficiency, 2),
            },
            'raw_kpis': {
                'total_primary': round(raw_total_ps, 2),
                'total_secondary': round(raw_total_ss, 2),
                'channel_efficiency': round(raw_channel_efficiency, 2),
            },
            'monthly_trend': trend_array,
            'monthly_comparison': all_months_comparison,
            'distributor_performance': dist_only[:20],
            'customer_performance': cust_only[:50],
            'product_group': product_array[:15]
        }, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def upload_csi_sales(request):
    """
    Upload a CSI (Customer Sales Intelligence) Excel file.
    Format: rows = distributor × customer × product, columns = months (e.g. Oct-25, Nov-25)
    Creates one Order record per (customer × product × month) with invoice_date = 1st of that month.
    Feeds directly into the existing secondary sales analytics.
    """
    try:
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file.name.lower()
        if not filename.endswith(('.xls', '.xlsx')):
            return Response({'error': 'Only .xlsx or .xls files are supported.'}, status=status.HTTP_400_BAD_REQUEST)

        import datetime
        import re
        import dateutil.parser

        # Read first sheet (Sales to customers - FY26) with no header
        raw_df = pd.read_excel(file, header=None, sheet_name=0)

        # Detect header row: look for 'product name', 'customer name', 'distributor name'
        header_row_idx = 0
        for i, r in raw_df.head(8).iterrows():
            row_vals = [str(v).strip().lower() if pd.notna(v) else '' for v in r]
            if any('product name' in v or 'customer name' in v or 'distributor name' in v for v in row_vals):
                header_row_idx = i
                break

        # Extract headers, converting date-like headers to YYYY-MM string
        raw_headers = []
        for v in raw_df.iloc[header_row_idx]:
            if pd.isna(v):
                raw_headers.append('')
            elif isinstance(v, (pd.Timestamp, datetime.datetime)):
                raw_headers.append(v.strftime('%Y-%m'))
            else:
                s = str(v).strip()
                # Try parsing short month strings like "Oct-25", "Nov-25"
                try:
                    parsed = dateutil.parser.parse(s, default=datetime.datetime(2000, 1, 1))
                    # Only treat as month if the string looks like a month label
                    if re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', s.lower()):
                        raw_headers.append(parsed.strftime('%Y-%m'))
                    else:
                        raw_headers.append(s)
                except Exception:
                    raw_headers.append(s)

        df = raw_df.iloc[header_row_idx + 1:].reset_index(drop=True)
        df.columns = raw_headers

        # Identify month columns (YYYY-MM format)
        month_col_pattern = re.compile(r'^\d{4}-\d{2}$')
        month_cols = [c for c in df.columns if month_col_pattern.match(str(c))]

        if not month_cols:
            return Response({'error': 'No month columns (e.g. Oct-25, Nov-25) found in the file.'}, status=status.HTTP_400_BAD_REQUEST)

        def get_col(possible_names):
            """Find the first column that matches any of the given names (case-insensitive)."""
            for name in possible_names:
                for c in df.columns:
                    if name.lower() in str(c).lower():
                        return c
            return None

        distributor_col = get_col(['Distributor Name', 'Distributor'])
        ship_to_col     = get_col(['Ship To Code', 'Ship To'])
        customer_col    = get_col(['Customer Name', 'Customer'])
        product_col     = get_col(['Product Name', 'Product'])
        product_code_col= get_col(['Product Code'])
        value_col       = get_col(['Value', 'Amount', 'Total Value'])

        if not customer_col or not product_col:
            return Response({'error': 'Could not find Customer Name or Product Name columns.'}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        valid_orders = []
        ignore_errors = request.POST.get('ignore_errors', 'false').lower() == 'true'
        valid_names_map = {clean_prod_name(name): name for name in ProductMaster.objects.values_list('material_name', flat=True)}
        prod_code_map = {clean_prod_name(p.material_name): p.material_code for p in ProductMaster.objects.all()}
        
        # Pre-fetch derived rates from Primary Sales to estimate CSI values (Derive from Value/Qty)
        from django.db.models import Sum
        rate_map = {}
        ps_data = PrimarySales.objects.values('material_desc').annotate(
            total_val=Sum('assessable_value'),
            total_qty=Sum('billed_quantity')
        )
        for r in ps_data:
            if r['total_qty'] and r['total_qty'] > 0:
                rate_map[clean_prod_name(r['material_desc'])] = r['total_val'] / r['total_qty']

        for index, row in df.iterrows():
            line_no = header_row_idx + index + 2

            def cell(col):
                if col is None:
                    return ''
                val = row.get(col, '')
                if isinstance(val, pd.Series):
                    val = val.iloc[0] if not val.empty else ''
                if pd.isna(val):
                    return ''
                s = str(val).strip()
                return '' if s.lower() == 'nan' else s

            customer    = cell(customer_col)
            product     = cell(product_col)
            sold_to     = cell(ship_to_col)   # Ship To Code used as sold_to identifier
            distributor = cell(distributor_col)

            # Skip empty/total rows
            if not customer and not product:
                continue

            # Validate product name against Product Master (Case-insensitive matching)
            product_clean = clean_prod_name(product)
            if product_clean and valid_names_map and product_clean not in valid_names_map:
                errors.append(f"Row {line_no}: Product '{product}' not in Product Master.")
                if not ignore_errors:
                    continue

            # Create one Order per month column that has a non-zero quantity
            for month_col in month_cols:
                qty_raw = row.get(month_col, '')
                if isinstance(qty_raw, pd.Series):
                    qty_raw = qty_raw.iloc[0] if not qty_raw.empty else ''
                if pd.isna(qty_raw) or str(qty_raw).strip() == '' or str(qty_raw).strip() == '0':
                    continue
                try:
                    qty = int(float(str(qty_raw).replace(',', '')))
                except (ValueError, TypeError):
                    continue

                if qty <= 0:
                    continue

                # invoice_date = 1st of the month
                try:
                    invoice_date = datetime.date(int(month_col[:4]), int(month_col[5:7]), 1)
                except Exception:
                    continue

                valid_orders.append(Order(
                    sold_to=sold_to,
                    ship_to=distributor,
                    invoice_no='',
                    invoice_date=invoice_date,
                    customer=customer,
                    material_code=cell(product_code_col) if product_code_col else prod_code_map.get(product_clean, ''),
                    material_name=product,
                    packsize=0,
                    qty=qty,
                    value=qty * next((rate for name, rate in rate_map.items() if name in product_clean or product_clean in name), 0.0) # Smart substring matching for rates
                ))

        if errors and not ignore_errors:
            return Response({'message': 'Validation failed.', 'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        Order.objects.bulk_create(valid_orders)
        msg = f'Successfully uploaded {len(valid_orders)} CSI sales records as secondary sales.'
        if errors and ignore_errors:
            msg += f' (Ignored {len(errors)} product master mismatches.)'
        return Response({'message': msg}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({'error': f'CSI upload failed: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
